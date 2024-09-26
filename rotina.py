import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font

# Criar os dados da rotina de estudo com horas líquidas e intervalos
data = {
    'Horário': [
        '04:30 – 06:30', '06:30 – 06:45', '06:45 – 08:45', '08:45 – 09:00', '09:00 – 15:00', 
        '15:00 – 16:30', '16:30 – 16:45', '16:45 – 17:45', '17:45 – 18:00', '18:00 – 21:40', 
        '21:40 – 22:30', '23:00 – Dormir'
    ],
    'Segunda a Sexta': [
        'Matemática: Álgebra', 'Intervalo', 'Física: Cinemática', 'Intervalo', 'Estágio', 
        'Química: Química Geral', 'Intervalo', 'Matemática: Cálculo', 'Intervalo', 'Faculdade', 
        'Revisão Geral ou Leitura (Português)', 'Dormir'
    ],
    'Sábado': [
        'Matemática: Geometria Analítica', 'Intervalo', 'Física: Dinâmica', 'Intervalo', 'Estágio', 
        'Química: Química Orgânica', 'Intervalo', 'Matemática: Matemática Discreta', 'Intervalo', 'Faculdade', 
        'Revisão: Termodinâmica e Óptica', 'Dormir'
    ],
    'Domingo': [
        'Física: Eletricidade e Magnetismo', 'Intervalo', 'Matemática: Estatística e Probabilidade', 'Intervalo', 'Química: Físico-Química', 
        'Português e Redação', 'Intervalo', 'Inglês', 'Intervalo', 'Revisão Geral', 
        'Revisão Leve', 'Dormir'
    ]
}

# Criar DataFrame
df = pd.DataFrame(data)

# Salvar em um arquivo Excel
file_path = "D:/OneDrive/Documentos/Camila/Github/rotina_estudo_ITA.xlsx"
df.to_excel(file_path, index=False)

# Abrir o arquivo Excel com openpyxl
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Adicionar uma nova planilha para gráficos
ws_summary = wb.create_sheet(title="Resumo")

# Adicionar dados de desempenho
# Exemplo: Coluna de estudo e notas
performance_data = {
    'Assunto': ['Matemática', 'Física', 'Química'],
    'Horas Estudadas': [10, 8, 7],  # Exemplo de dados
    'Desempenho (%)': [85, 90, 75]  # Exemplo de dados
}

df_performance = pd.DataFrame(performance_data)

# Adicionar dados ao novo sheet
for row in dataframe_to_rows(df_performance, index=False, header=True):
    ws_summary.append(row)

# Criar gráfico de barras para horas estudadas
chart_hours = BarChart()
chart_hours.title = "Horas Estudadas"
chart_hours.x_axis.title = "Assunto"
chart_hours.y_axis.title = "Horas"
data_hours = Reference(ws_summary, min_col=2, min_row=1, max_col=2, max_row=len(df_performance) + 1)
categories = Reference(ws_summary, min_col=1, min_row=2, max_row=len(df_performance) + 1)
chart_hours.add_data(data_hours, titles_from_data=True)
chart_hours.set_categories(categories)
ws_summary.add_chart(chart_hours, "E5")

# Criar gráfico de barras para desempenho
chart_performance = BarChart()
chart_performance.title = "Desempenho"
chart_performance.x_axis.title = "Assunto"
chart_performance.y_axis.title = "Desempenho (%)"
data_performance = Reference(ws_summary, min_col=3, min_row=1, max_col=3, max_row=len(df_performance) + 1)
chart_performance.add_data(data_performance, titles_from_data=True)
chart_performance.set_categories(categories)
ws_summary.add_chart(chart_performance, "E20")

# Adicionar formatação em tons de rosa
pink_fill = PatternFill(start_color="F4C6C6", end_color="F4C6C6", fill_type="solid")
font = Font(color="C810F0")

for row in ws_summary.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
    for cell in row:
        cell.fill = pink_fill
        cell.font = font

# Salvar o arquivo
wb.save(file_path)
print(f"Planilha com gráficos salva como {file_path}")
